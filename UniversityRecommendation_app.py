import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import os
import re
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
import hashlib
import time
import json

# 페이지 설정
st.set_page_config(
    page_title="코드스튜디오 입시연구소",
    page_icon="🎓",
    layout="wide"
)

# 세션 상태 초기화
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.user = None
    st.session_state.license_key = None

# 라이센스 체크 함수
def check_license():
    """라이센스 확인"""
    try:
        licenses = st.secrets["licenses"]
        return licenses
    except KeyError:
        return None

def get_gsheet_client():
    """Google Sheets 클라이언트 생성"""
    try:
        # secrets에서 credentials 정보 가져오기
        creds_info = dict(st.secrets["gsheets"]["credentials"])
        
        # private_key의 줄바꿈 처리
        if 'private_key' in creds_info:
            creds_info['private_key'] = creds_info['private_key'].replace('\\n', '\n')
        
        # Credentials 객체 생성
        creds = Credentials.from_service_account_info(
            creds_info,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        
        # gspread 클라이언트 생성
        client = gspread.authorize(creds)
        return client
        
    except Exception as e:
        st.error(f"Google Sheets 클라이언트 생성 실패: {str(e)}")
        return None

def test_google_sheets_connection():
    """Google Sheets 연결 테스트 함수"""
    try:
        client = get_gsheet_client()
        if not client:
            return False, "클라이언트 생성 실패"
        
        # 스프레드시트 ID 확인
        spreadsheet_id = st.secrets.get("gsheets", {}).get("spreadsheet_id")
        if not spreadsheet_id:
            return False, "스프레드시트 ID가 설정되지 않았습니다"
        
        # 스프레드시트 열기 시도
        sheet = client.open_by_key(spreadsheet_id)
        worksheet = sheet.get_worksheet(0)  # 첫 번째 워크시트
        
        # 테스트로 A1 셀 읽기
        test_value = worksheet.get('A1')
        
        return True, "연결 성공"
        
    except gspread.exceptions.SpreadsheetNotFound:
        return False, "스프레드시트를 찾을 수 없습니다. ID를 확인하세요."
    except gspread.exceptions.APIError as e:
        if e.response.status_code == 403:
            return False, "권한이 없습니다. 서비스 계정에 스프레드시트 접근 권한을 부여하세요."
        elif e.response.status_code == 404:
            return False, "스프레드시트를 찾을 수 없습니다. ID를 확인하세요."
        else:
            return False, f"API 오류: {e.response.status_code} - {e.response.text}"
    except Exception as e:
        return False, f"알 수 없는 오류: {str(e)}"

def log_user_activity(user, activity_type="login"):
    """사용자 활동 로그 기록 - 실패해도 앱은 계속 실행"""
    try:
        client = get_gsheet_client()
        if not client:
            # 로그만 남기고 계속 진행
            return False
            
        spreadsheet_id = st.secrets.get("gsheets", {}).get("spreadsheet_id")
        if not spreadsheet_id:
            return False
            
        # 스프레드시트 열기
        sheet = client.open_by_key(spreadsheet_id)
        
        # 로그 워크시트 찾기 또는 생성
        try:
            log_sheet = sheet.worksheet("사용자로그")
        except:
            # 워크시트가 없으면 생성
            log_sheet = sheet.add_worksheet(title="사용자로그", rows=1000, cols=10)
            # 헤더 추가
            log_sheet.update('A1:E1', [['사용자', '활동유형', '시간', 'IP', '세부정보']])
        
        # 새 로그 추가
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_row = [user, activity_type, timestamp, "", ""]
        
        # 다음 빈 행에 추가
        log_sheet.append_row(new_row)
        
        return True
        
    except Exception as e:
        # 오류가 나도 앱은 계속 실행
        return False

# 라이센스 인증 화면
if not st.session_state.authenticated:
    st.title("🎓 코드스튜디오 입시연구소")
    st.markdown("### 라이센스 인증이 필요합니다")
    
    licenses = check_license()
    if licenses is None:
        st.error("시스템 설정 오류: 관리자에게 문의하세요.")
        st.stop()
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        license_key = st.text_input("라이센스 키를 입력하세요", 
                                   type="password",
                                   placeholder="예: RFKX-ZWWU-860D-A8MO")
        
        if st.button("확인", use_container_width=True, type="primary"):
            if license_key:
                # 라이센스 검증
                valid = False
                for license in licenses:
                    if license["key"] == license_key:
                        st.session_state.authenticated = True
                        st.session_state.user = license["user"]
                        st.session_state.license_key = license_key
                        valid = True
                        
                        # 로그인 로그 기록 시도 (실패해도 계속 진행)
                        try:
                            log_user_activity(license["user"], "login")
                        except Exception as e:
                            # 로그 실패해도 계속 진행
                            pass
                        
                        st.success(f"✅ 환영합니다, {license['user']}님!")
                        st.balloons()
                        st.rerun()
                        break
                
                if not valid:
                    st.error("❌ 유효하지 않은 라이센스 키입니다.")
            else:
                st.warning("⚠️ 라이센스 키를 입력해주세요.")
    
    with st.expander("도움말"):
        st.markdown("""
        - 라이센스 키는 XXXX-XXXX-XXXX-XXXX 형식입니다
        - 대소문자를 정확히 입력해주세요
        - 문제가 있으면 관리자에게 문의하세요
        """)
    st.stop()

# ===== 여기서부터는 메인 화면 =====

# 제목
st.title("🎓 코드스튜디오 입시연구소")
st.markdown("### 2021~2025년 5개년 데이터 기반 맞춤 추천")

# 우측 상단에 로그아웃 버튼
col1, col2 = st.columns([10, 1])

# Google Sheets 연결 테스트 (관리자용)
if st.checkbox("🔧 시스템 상태 확인 (관리자용)"):
    st.subheader("Google Sheets 연결 상태")
    
    # 연결 테스트
    success, message = test_google_sheets_connection()
    
    if success:
        st.success(f"✅ {message}")
        
        # 추가 정보 표시
        try:
            client = get_gsheet_client()
            spreadsheet_id = st.secrets.get("gsheets", {}).get("spreadsheet_id")
            sheet = client.open_by_key(spreadsheet_id)
            
            st.info(f"""
            **연결 정보:**
            - 스프레드시트 ID: {spreadsheet_id}
            - 워크시트 수: {len(sheet.worksheets())}
            - 서비스 계정: {st.secrets["gsheets"]["credentials"]["client_email"]}
            """)
            
            # 워크시트 목록
            st.write("**워크시트 목록:**")
            for ws in sheet.worksheets():
                st.write(f"- {ws.title}")
                
        except Exception as e:
            st.error(f"상세 정보 조회 실패: {str(e)}")
    else:
        st.error(f"❌ {message}")
        
        # 해결 방법 안내
        st.info("""
        **해결 방법:**
        1. Google Sheets에서 스프레드시트를 열어주세요
        2. 공유 버튼 클릭
        3. 다음 이메일 추가: `google-sheets-api@stoked-name-475406-h9.iam.gserviceaccount.com`
        4. '편집자' 권한 부여
        5. 완료 후 페이지 새로고침
        """)

with col2:
    if st.button("로그아웃"):
        # 로그아웃 로그 기록 시도 (실패해도 계속 진행)
        try:
            log_user_activity(st.session_state.user, "logout")
        except:
            pass
        
        st.session_state.authenticated = False
        st.session_state.user = None
        st.session_state.license_key = None
        st.rerun()

st.markdown("---")

# CSV 데이터 로드
@st.cache_data
def load_admissions_data():
    """입시 데이터 CSV 로드"""
    import chardet
    
    file_path = '2025_2021_result.csv'
    
    # 파일 존재 확인
    if not os.path.exists(file_path):
        st.error(f"CSV 파일을 찾을 수 없습니다: {file_path}")
        return None
    
    # 인코딩 감지
    try:
        with open(file_path, 'rb') as f:
            raw_data = f.read(100000)
            result = chardet.detect(raw_data)
            detected_encoding = result['encoding']
            st.sidebar.info(f"감지된 인코딩: {detected_encoding}")
    except:
        detected_encoding = None
    
    # 여러 인코딩으로 시도
    encodings = [detected_encoding, 'utf-8-sig', 'utf-8', 'cp949', 'euc-kr', 'latin1']
    encodings = [e for e in encodings if e]
    
    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, encoding=encoding)
            
            # 데이터 검증
            if len(df.columns) == 13 and len(df) > 0:
                # 컬럼명 설정
                df.columns = [
                    'year', 'university_name', 'admission_type', 'admission_name',
                    'major_name', 'quota', 'comp_rate', 'pass_rank',
                    'cut_grade_50', 'cut_grade_70', 'cut_grade_85', 'cut_grade_90',
                    'reflected_subjects'
                ]
                
                # 데이터 타입 변환
                numeric_cols = ['quota', 'comp_rate', 'pass_rank', 
                               'cut_grade_50', 'cut_grade_70', 'cut_grade_85', 'cut_grade_90']
                for col in numeric_cols:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                
                st.sidebar.success(f"✅ CSV 로드 성공 (인코딩: {encoding})")
                st.sidebar.write(f"데이터 수: {len(df):,}개")
                
                return df
                
        except Exception as e:
            continue
    
    # 파일 업로드 옵션 제공
    st.sidebar.error("자동 로드 실패. 파일을 직접 업로드해주세요.")
    uploaded_file = st.sidebar.file_uploader(
        "CSV 파일 직접 업로드",
        type=['csv'],
        help="2025_2021_result.csv 파일을 선택해주세요"
    )
    
    if uploaded_file:
        try:
            df = pd.read_csv(uploaded_file, encoding='utf-8-sig')
            
            if len(df.columns) == 13:
                df.columns = [
                    'year', 'university_name', 'admission_type', 'admission_name',
                    'major_name', 'quota', 'comp_rate', 'pass_rank',
                    'cut_grade_50', 'cut_grade_70', 'cut_grade_85', 'cut_grade_90',
                    'reflected_subjects'
                ]
                
                st.sidebar.success("✅ 업로드 파일 로드 성공!")
                return df
        except Exception as e:
            st.sidebar.error(f"업로드 파일 오류: {str(e)}")
    
    return None

def read_student_info_from_excel(excel_file):
    """내신분석 시트에서 학생 정보 추출"""
    try:
        wb = load_workbook(excel_file, data_only=True)
        
        st.info(f"📋 엑셀 시트 목록: {wb.sheetnames}")
        
        # Index 시트에서 정보 추출
        if 'Index' in wb.sheetnames:
            ws = wb['Index']
            
            # 학교명: F4, F5
            school_name = ws['F4'].value or ws['F5'].value
            
            # 학년: I4, I5
            grade = ws['I4'].value or ws['I5'].value
            
            # 이름: K4, K5, L4, L5
            student_name = (ws['K4'].value or ws['K5'].value or 
                          ws['L4'].value or ws['L5'].value)
            
            wb.close()
            
            # 학년 처리
            if grade:
                grade_str = str(grade).strip()
                numbers = re.findall(r'\d+', grade_str)
                if numbers:
                    grade = f"{numbers[0]}학년"
                elif '학년' in grade_str:
                    grade = grade_str
                else:
                    grade = f"{grade_str}학년"
            else:
                grade = "2학년"
            
            result = {
                'name': str(student_name).strip() if student_name else '',
                'school': str(school_name).strip() if school_name else '',
                'grade': grade
            }
            
            st.success(f"✅ 추출된 학생 정보: {result}")
            return result
        else:
            st.error("'Index' 시트를 찾을 수 없습니다.")
            wb.close()
            return None
    except Exception as e:
        st.error(f"❌ 학생 정보 추출 오류: {str(e)}")
        return None

def get_student_grade_from_excel(excel_file):
    """성적분석 시트의 X13 셀에서 평균 등급 추출"""
    try:
        wb = load_workbook(excel_file, data_only=True)
        
        if '성적분석' in wb.sheetnames:
            ws = wb['성적분석']
            
            # X13 셀에서 전과목 평균 읽기
            avg_grade = ws['X13'].value
            
            wb.close()
            
            if avg_grade and isinstance(avg_grade, (int, float)):
                st.success(f"✅ 전과목 평균: {avg_grade}등급")
                return float(avg_grade)
        
        wb.close()
        return 2.5
    except Exception as e:
        st.warning(f"성적 자동 추출 실패: {str(e)}")
        return 2.5

def get_major_keywords(df):
    """학과명에서 핵심 단어 추출"""
    if df is None or 'major_name' not in df.columns:
        return []
    
    all_majors = df['major_name'].dropna().unique()
    
    # 키워드 빈도 계산
    keyword_freq = {}
    
    # 제외할 단어들
    exclude_words = {
        '학과', '과', '전공', '부', '학부', '계열', '및', '와', '의', 
        '(', ')', '・', ',', '-', '/', ' ', '전공학'
    }
    
    for major in all_majors:
        major = str(major)
        
        # 괄호 안 내용 제거
        major = re.sub(r'\([^)]*\)', '', major)
        
        # 여러 구분자로 단어 분리
        words = re.split(r'[(\s)・,/-]+', major)
        
        for word in words:
            word = word.strip()
            
            # 2글자 이상, 제외 단어 아님
            if len(word) >= 2 and word not in exclude_words:
                keyword_freq[word] = keyword_freq.get(word, 0) + 1
    
    # 빈도수 순으로 정렬
    popular_keywords = [
        k for k, v in sorted(keyword_freq.items(), key=lambda x: x[1], reverse=True) 
        if len(k) >= 2
    ]
    
    return popular_keywords[:300]

def flexible_search(text, keyword):
    """유연한 검색"""
    if pd.isna(text) or not keyword:
        return False
    
    text = str(text).lower()
    keyword = str(keyword).lower()
    
    # 공백으로 구분된 여러 키워드
    keywords = keyword.split()
    
    if not keywords:
        return False
    
    # 각 키워드에 대해 검사
    for kw in keywords:
        kw = kw.strip()
        if not kw:
            continue
            
        # 공백 제거하여 검색
        text_no_space = text.replace(' ', '').replace('・', '')
        kw_no_space = kw.replace(' ', '').replace('・', '')
        
        # 키워드가 텍스트에 포함되어 있으면 True
        if kw in text or kw_no_space in text_no_space:
            return True
    
    return False

def categorize_university(student_grade, cut_grade):
    """대학을 구분별로 분류"""
    # 학생 등급 - 합격선 등급
    diff = student_grade - cut_grade
    
    if diff >= 1.5:
        return '강상향'
    elif diff >= 0.8:
        return '상향'
    elif diff >= 0.3:
        return '약상향'
    elif diff >= -0.3:
        return '적정'
    elif diff >= -0.8:
        return '강적정'
    elif diff >= -1.5:
        return '안정'
    else:
        return '강안정'

def get_category_color(category):
    """구분별 색상"""
    colors = {
        '강상향': '#ef4444',
        '상향': '#f97316',
        '약상향': '#eab308',
        '적정': '#22c55e',
        '강적정': '#10b981',
        '안정': '#3b82f6',
        '강안정': '#6366f1',
        '정보없음': '#9ca3af'
    }
    return colors.get(category, '#6b7280')

def find_recommendations(df, major_keyword, student_grade, num_results=30):
    """대학 추천"""
    
    # 유연한 검색 적용
    filtered = df[df['major_name'].apply(lambda x: flexible_search(x, major_keyword))]
    
    if len(filtered) == 0:
        return None, None, f"'{major_keyword}' 관련 학과를 찾을 수 없습니다."
    
    # 년도별 가중치 설정
    year_weights = {
        '2025': 1.0,
        '2024': 0.8,
        '2023': 0.6,
        '2022': 0.4,
        '2021': 0.3
    }
    
    results = []
    category_distribution = {}
    
    # 대학-학과별로 그룹화
    grouped = filtered.groupby(['university_name', 'major_name', 'admission_type', 'admission_name'])
    
    for (univ, major, adm_type, adm_name), group in grouped:
        # 가중평균 계산을 위한 변수
        weighted_cuts = []
        weights_sum = 0
        comp_rates = []
        
        # 최신 데이터 가져오기
        latest_row = group[group['year'] == group['year'].max()].iloc[0]
        latest_cut_70 = None
        
        # 최신 년도의 70%컷 찾기
        if pd.notna(latest_row['cut_grade_70']) and latest_row['cut_grade_70'] > 0:
            latest_cut_70 = float(latest_row['cut_grade_70'])
        
        for _, row in group.iterrows():
            year = str(row.get('year', '2025'))
            weight = year_weights.get(year, 0.5)
            
            # 여러 컷 중 하나 선택
            cut_grade = None
            for col in ['cut_grade_70', 'cut_grade_50', 'cut_grade_85', 'cut_grade_90']:
                if pd.notna(row[col]) and row[col] > 0:
                    try:
                        cut_grade = float(row[col])
                        break
                    except:
                        continue
            
            if cut_grade:
                weighted_cuts.append(cut_grade * weight)
                weights_sum += weight
            
            # 경쟁률 수집
            if pd.notna(row['comp_rate']):
                comp_rates.append(float(row['comp_rate']))
        
        # 가중평균 컷라인 계산
        if weights_sum > 0:
            avg_cut_grade = sum(weighted_cuts) / weights_sum
        else:
            avg_cut_grade = None
        
        # 종합전형 여부 확인
        is_jonghap = '종합' in str(adm_type)
        
        # 카테고리 분류
        if avg_cut_grade and avg_cut_grade > 0:
            category = categorize_university(float(student_grade), avg_cut_grade)
            diff = abs(float(student_grade) - avg_cut_grade)
        else:
            category = '정보없음'
            diff = 999
        
        # 안정성 계산
        if len(weighted_cuts) > 1:
            grades = [c/w for c, w in zip(weighted_cuts, [year_weights.get(str(row.get('year', '2025')), 0.5) for _, row in group.iterrows()]) if w > 0]
            if len(grades) > 1:
                stability = np.std(grades)
            else:
                stability = 0
        else:
            stability = 999
        
        # 평균 경쟁률
        avg_comp_rate = np.mean(comp_rates) if comp_rates else latest_row.get('comp_rate', None)
        
        # 카테고리별 개수 세기
        category_distribution[category] = category_distribution.get(category, 0) + 1
        
        results.append({
            'university': univ,
            'major': major,
            'admission_type': adm_type,
            'admission_name': adm_name,
            'category': category,
            'diff': diff,
            'cut_grade': avg_cut_grade,
            'comp_rate': avg_comp_rate,
            'is_jonghap': is_jonghap,
            'priority': 0 if is_jonghap else 1,
            'stability': stability,
            'years_data': len(group),
            'latest_cut_70': latest_cut_70
        })
    
    # 구분별 분포 표시
    with st.expander("📊 구분별 학과 분포"):
        jonghap_count = sum(1 for r in results if r['is_jonghap'])
        st.write(f"**종합전형**: {jonghap_count}개 | **교과전형**: {len(results) - jonghap_count}개")
        st.write("---")
        for cat in ['강상향', '상향', '약상향', '적정', '강적정', '안정', '강안정', '정보없음']:
            count = category_distribution.get(cat, 0)
            if count > 0:
                st.write(f"**{cat}**: {count}개")
    
    # 추천 전략
    recommendations = []
    used = set()
    
    # 카테고리별 목표 개수
    category_targets = {
        '강상향': 3,
        '상향': 5,
        '약상향': 5,
        '적정': 7,
        '강적정': 5,
        '안정': 3,
        '강안정': 2
    }
    
    # 각 구분별로 목표 개수만큼 선택
    for cat, target_count in category_targets.items():
        cat_results = [r for r in results if r['category'] == cat]
        if cat_results:
            sorted_results = sorted(cat_results, 
                key=lambda x: (x['priority'], x['diff'], x['stability']))
            
            added = 0
            for result in sorted_results:
                if added >= target_count:
                    break
                key = (result['university'], result['major'])
                if key not in used:
                    recommendations.append(result)
                    used.add(key)
                    added += 1
    
    # 30개가 안 되면 추가
    if len(recommendations) < num_results:
        remaining = [r for r in results if (r['university'], r['major']) not in used]
        sorted_remaining = sorted(remaining, 
            key=lambda x: (x['priority'], x['diff'], x['stability']))
        
        for result in sorted_remaining:
            if len(recommendations) >= num_results:
                break
            key = (result['university'], result['major'])
            if key not in used:
                recommendations.append(result)
                used.add(key)
    
    return recommendations[:num_results], filtered, None

def create_excel_output(student_info, recommendations, all_results_df=None):
    """엑셀 파일 생성"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # 첫 번째 시트: 학교추천
    ws1 = wb.active
    ws1.title = "학교추천"
    
    # 스타일 정의
    orange_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF", size=11)
    black_font = Font(size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 1행 - 학교 정보 헤더
    headers_row1 = [
        ('A1', '학교명', orange_fill, white_font),
        ('C1', '학년', orange_fill, white_font),
        ('E1', '이름', orange_fill, white_font),
        ('G1', '희망진로', orange_fill, white_font)
    ]
    
    for cell_addr, value, fill, font in headers_row1:
        cell = ws1[cell_addr]
        cell.value = value
        cell.fill = fill
        cell.font = font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 1행 - 학교 정보 데이터
    ws1['B1'] = student_info['school']
    ws1['D1'] = student_info['grade']
    ws1['F1'] = student_info['name']
    ws1['H1'] = student_info['major']
    
    # 병합할 셀들
    ws1.merge_cells('H1:J1')
    
    # 3행 - 테이블 헤더
    headers_row3 = ['학교', '학과명', '전형', '전형요소', '구분', '최근70%컷', '데이터년수', '평균경쟁률']
    for idx, header in enumerate(headers_row3, start=1):
        cell = ws1.cell(row=3, column=idx)
        cell.value = header
        cell.fill = orange_fill
        cell.font = white_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # 4행부터 데이터 입력
    for idx, rec in enumerate(recommendations, start=4):
        ws1[f'A{idx}'] = rec['university']
        ws1[f'B{idx}'] = rec['major']
        ws1[f'C{idx}'] = rec['admission_type']
        ws1[f'D{idx}'] = rec['admission_name']
        ws1[f'E{idx}'] = rec['category']
        ws1[f'F{idx}'] = f"{rec.get('latest_cut_70', '-'):.2f}" if rec.get('latest_cut_70') and rec.get('latest_cut_70') != 999 else "-"
        ws1[f'G{idx}'] = f"{rec.get('years_data', 1)}년"
        ws1[f'H{idx}'] = f"{rec.get('comp_rate', '-'):.1f}" if rec.get('comp_rate') else "-"
        
        # 모든 셀에 테두리와 정렬 적용
        for col in range(1, 9):
            cell = ws1.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = center_align
            cell.font = black_font
        
        # 구분 셀 색상 적용
        category_colors = {
            '강상향': 'FF9999', '상향': 'FFB366', '약상향': 'FFCC66',
            '적정': '99FF99', '강적정': '66FFB3', '안정': '99CCFF', 
            '강안정': '9999FF', '정보없음': 'E6E6E6'
        }
        
        color = category_colors.get(rec['category'], 'FFFFFF')
        ws1[f'E{idx}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # 평균경쟁률 셀 색상
        ws1[f'H{idx}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # 열 너비 조정
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 15
    
    # 두 번째 시트: 전체 검색 결과
    if all_results_df is not None:
        try:
            ws2 = wb.create_sheet("전체검색결과")
            
            # 헤더
            headers = ['년도', '대학명', '학과명', '전형', '경쟁률', '50%컷', '70%컷']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws2.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = orange_fill
                cell.font = white_font
                cell.alignment = center_align
                cell.border = thin_border
            
            # 데이터 입력
            row_idx = 2
            for _, row in all_results_df.iterrows():
                ws2.cell(row=row_idx, column=1).value = str(row['year']) if pd.notna(row['year']) else ''
                ws2.cell(row=row_idx, column=2).value = str(row['university_name']) if pd.notna(row['university_name']) else ''
                ws2.cell(row=row_idx, column=3).value = str(row['major_name']) if pd.notna(row['major_name']) else ''
                ws2.cell(row=row_idx, column=4).value = str(row['admission_type']) if pd.notna(row['admission_type']) else ''
                ws2.cell(row=row_idx, column=5).value = row['comp_rate'] if pd.notna(row['comp_rate']) else '-'
                ws2.cell(row=row_idx, column=6).value = row['cut_grade_50'] if pd.notna(row['cut_grade_50']) else '-'
                ws2.cell(row=row_idx, column=7).value = row['cut_grade_70'] if pd.notna(row['cut_grade_70']) else '-'
                
                row_idx += 1
            
            # 열 너비 조정
            ws2.column_dimensions['A'].width = 10
            ws2.column_dimensions['B'].width = 20
            ws2.column_dimensions['C'].width = 35
            ws2.column_dimensions['D'].width = 15
            ws2.column_dimensions['E'].width = 10
            ws2.column_dimensions['F'].width = 10
            ws2.column_dimensions['G'].width = 10
        except Exception as e:
            st.warning(f"전체 검색 결과 시트 생성 중 오류: {str(e)}")
    
    # BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

# 메인 애플리케이션
def main():
    with st.sidebar:
        st.header("📚 시스템 정보")
        st.info("5개년 데이터 기반 30개 대학 추천")
        st.write(f"**사용자**: {st.session_state.user}")
    
    df = load_admissions_data()
    
    if df is None:
        st.error("⚠️ CSV 파일을 로드할 수 없습니다.")
        
        # CSV 파일 생성 도움말
        with st.expander("CSV 파일이 없는 경우"):
            st.markdown("""
            **CSV 파일 구조 (13개 컬럼):**
            1. 년도
            2. 대학명
            3. 중심전형
            4. 전형명
            5. 모집단위
            6. 모집인원
            7. 경쟁률
            8. 충원순위
            9. 50%컷
            10. 70%컷
            11. 85%컷
            12. 90%컷
            13. 반영교과목
            
            파일명: `2025_2021_result.csv`
            """)
        st.stop()
    
    st.success(f"✅ 입시 데이터: {len(df):,}개 (2021~2025)")
    
    # 데이터 통계 정보
    with st.expander("📊 데이터 상세 정보"):
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("총 데이터", f"{len(df):,}개")
            year_stats = df['year'].value_counts().sort_index()
            for year, count in year_stats.items():
                st.write(f"{year}년: {count:,}개")
        with col2:
            st.metric("대학 수", f"{df['university_name'].nunique():,}개")
            st.metric("학과 수", f"{df['major_name'].nunique():,}개")
        with col3:
            st.metric("종합전형", f"{df[df['admission_type'].str.contains('종합', na=False)].shape[0]:,}개")
            st.metric("교과전형", f"{df[df['admission_type'].str.contains('교과', na=False)].shape[0]:,}개")
    
    major_keywords = get_major_keywords(df)
    st.sidebar.info(f"✅ {len(major_keywords)}개의 학과 키워드 추출 완료")
    
    st.subheader("📄 1. 평가표 업로드")
    uploaded_file = st.file_uploader(
        "엑셀 파일을 업로드하세요",
        type=['xlsx', 'xls']
    )
    
    student_name = ""
    school_name = ""
    grade = "2학년"
    student_grade = 2.5
    
    if uploaded_file:
        student_info = read_student_info_from_excel(uploaded_file)
        
        if student_info:
            if student_info['name']:
                student_name = student_info['name']
            if student_info['school']:
                school_name = student_info['school']
            if student_info['grade']:
                grade = student_info['grade']
        
        uploaded_file.seek(0)
        auto_grade = get_student_grade_from_excel(uploaded_file)
        if auto_grade:
            student_grade = auto_grade
    
    st.subheader("👤 2. 학생 정보")
    st.info("💡 엑셀에서 자동 추출된 정보입니다. 비어있으면 직접 입력해주세요.")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        student_name = st.text_input("이름", value=student_name, placeholder="홍길동")
    with col2:
        school_name = st.text_input("학교명", value=school_name, placeholder="코드고등학교")
    with col3:
        grade_options = ["1학년", "2학년", "3학년"]
        grade_index = grade_options.index(grade) if grade in grade_options else 1
        grade = st.selectbox("학년", grade_options, index=grade_index)
    
    st.subheader("📊 3. 내신 성적")
    student_grade = st.number_input("내신 평균 등급", 1.0, 9.0, float(student_grade), 0.1)
    
    st.subheader("🎯 4. 희망 전공")
    
    search_mode = st.radio("검색 방식", ["직접 입력", "키워드 선택"], horizontal=True)
    
    if search_mode == "키워드 선택":
        hope_major = st.selectbox("학과 키워드", [""] + major_keywords[:100])
    else:
        hope_major = st.text_input("키워드 입력", placeholder="예: 컴퓨터, 기계, 전자")
    
    if hope_major:
        matching = df[df['major_name'].apply(lambda x: flexible_search(x, hope_major))]
        unique_majors = matching.groupby(['university_name', 'major_name']).size().reset_index()
        st.metric("매칭 학과", f"{len(unique_majors)}개 대학/학과")
    
    st.markdown("---")
    
    if st.button("🚀 대학 추천 실행 (30개)", type="primary", use_container_width=True):
        if not student_name or not school_name or not hope_major:
            st.error("모든 정보를 입력해주세요.")
        else:
            with st.spinner("5개년 데이터를 분석하여 추천 중..."):
                # 추천 활동 로그 시도 (실패해도 계속 진행)
                try:
                    log_user_activity(st.session_state.user, f"recommend_{hope_major}")
                except:
                    pass
                
                recommendations, filtered, error = find_recommendations(df, hope_major, student_grade)
                
                if error:
                    st.error(error)
                else:
                    st.success(f"✅ {len(recommendations)}개 대학 추천 완료!")
                    
                    st.session_state['recommendations'] = recommendations
                    st.session_state['student_info'] = {
                        'name': student_name,
                        'school': school_name,
                        'grade': grade,
                        'major': hope_major
                    }
                    st.session_state['filtered_df'] = filtered
                    
                    # 결과 표시
                    df_results = pd.DataFrame(recommendations)
                    display_df = df_results[['category', 'university', 'major', 'admission_type', 
                                           'latest_cut_70', 'cut_grade', 'comp_rate', 'years_data']].copy()
                    display_df.columns = ['구분', '대학명', '학과명', '전형', '최근70%컷', '평균합격선', '평균경쟁률', '데이터년수']
                    
                    # 포맷팅
                    display_df['최근70%컷'] = display_df['최근70%컷'].apply(lambda x: f"{x:.2f}" if pd.notna(x) and x != 999 else "-")
                    display_df['평균합격선'] = display_df['평균합격선'].apply(lambda x: f"{x:.2f}" if pd.notna(x) and x != 999 else "-")
                    display_df['평균경쟁률'] = display_df['평균경쟁률'].apply(lambda x: f"{x:.1f}" if pd.notna(x) else "-")
                    display_df['데이터년수'] = display_df['데이터년수'].apply(lambda x: f"{x}년")
                    
                    # 스타일 적용
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        height=600
                    )
    
    if 'recommendations' in st.session_state:
        st.markdown("---")
        
        # 추천 통계 표시
        with st.expander("📈 추천 결과 통계"):
            recs = st.session_state['recommendations']
            col1, col2, col3 = st.columns(3)
            
            with col1:
                category_counts = pd.Series([r['category'] for r in recs]).value_counts()
                st.write("**구분별 분포:**")
                for cat, count in category_counts.items():
                    st.write(f"- {cat}: {count}개")
            
            with col2:
                jonghap_count = sum(1 for r in recs if r['is_jonghap'])
                st.write("**전형별 분포:**")
                st.write(f"- 종합전형: {jonghap_count}개")
                st.write(f"- 교과전형: {len(recs) - jonghap_count}개")
            
            with col3:
                year_counts = pd.Series([r['years_data'] for r in recs]).value_counts().sort_index()
                st.write("**데이터 년수별:**")
                for years, count in year_counts.items():
                    st.write(f"- {years}년 데이터: {count}개")
        
        # 엑셀 다운로드
        output_file = create_excel_output(
            st.session_state['student_info'],
            st.session_state['recommendations'],
            st.session_state.get('filtered_df', None)
        )
        
        # 다운로드 로그 시도 (실패해도 계속 진행)
        if st.download_button(
            "📥 엑셀 파일 다운로드",
            output_file,
            f"대학추천_{st.session_state['student_info']['name']}_{len(st.session_state['recommendations'])}개.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        ):
            try:
                log_user_activity(st.session_state.user, "download_excel")
            except:
                pass

if __name__ == "__main__":
    main()
